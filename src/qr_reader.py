from tkinter import *
from datetime import datetime
from PIL import Image, ImageTk
from threading import Timer
import tkinter as tk

import time
import sys
import openpyxl

def clear_frame(frame):
	for widgets in frame.winfo_children():
		widgets.pack_forget()


## Choose Service ##
service_num = int(input("Which Service(1-3): "))
while service_num <= 0 | service_num >= 4:
    print("Enter between 1 - 3.")
    service_num = int(input("Which service(1-3): "))

## set time ##
now = datetime.now()
month_list = {1:"Jan ", 2:"Feb ", 3:"Mar ", 4:"Apr ", 5:"May ", 6:"Jun ",
              7:"Jul ", 8:"Aug ", 9:"Sep ", 10:"Oct ", 11:"Nov ", 12:"Dec "}
service_list = {1:" at 8:00 AM", 2:" at 10:30 AM", 3:" at 1:00 PM"}

month = now.month
date = str(now.day)
year = str(now.year)

this_service = month_list[month] + date + ", " + year + service_list[service_num]
print("Today's Service: " + this_service)

## database path ##
# EventBrite List #
loc  = "data/report-2020-11-10T1105.xlsx"
# Pastor / Volunteers
loc2 = ("data/11082020_list.xlsx")

# input file prep #
wb = openpyxl.load_workbook(loc)
sheet = wb.active

# pastor/volunteer list loading
wb = openpyxl.load_workbook(loc2)
pSheet = wb.active

# format checking #
if sheet['J1'].value != "Barcode #":
    print("ERROR: Barcode # is not on excel file")
    quit()

# output file prep #
logfile = ("log/" + now.strftime("%Y_%m_%d__%H_%M_%S")+".log")
f = open(logfile, "x", encoding='utf-8')  # create file for writing
f.write(this_service + "\n")

# read-up ticket and makes dictionary #
attendee_list = {
        "barcode#": ["name", "email", "phone", "checkin"]
}

entered_person = {
        "barcode#": "Time"
}

## insert data to database ##
for i in range(2, sheet.max_row + 1):
    first_name  = sheet.cell(i, 5).value
    last_name   = sheet.cell(i, 6).value
    name = first_name + " " + last_name
    email = sheet.cell(i,7).value
    barcode_num = sheet.cell(i,10).value
    checkIn = sheet.cell(i, 11).value
    phone = sheet.cell(i,13).value
    attendee_list[barcode_num] = [name, email, phone, checkIn]

# Pastors / Volunteers
for i in range(2, pSheet.max_row + 1):
    pName = pSheet.cell(i, 2).value
    barcode_num = pSheet.cell(i, 1).value
    pEmail = pSheet.cell(i, 3).value
    if(pSheet.cell(i, 4).value != None):
        pPhone = pSheet.cell(i, 4).value
    else:
        pPhone = ""
    pCheckin = this_service
    attendee_list[barcode_num] = [pName, pEmail, pPhone, pCheckin]

# print(attendee_list)

## GUIs ##
ROOT = Tk()
ROOT.title("EPC Sunday Service")
ROOT.tk_setPalette(background='white')
ROOT.geometry('1680x800')
frame = Frame(ROOT);
frame.pack();

## images ##
wel_img1 = ImageTk.PhotoImage(Image.open("img/welcome1.png"))
wel_img2 = ImageTk.PhotoImage(Image.open("img/welcome2.png"))
wel_img3 = ImageTk.PhotoImage(Image.open("img/welcome3.png"))
qr_img = ImageTk.PhotoImage(Image.open("img/qr_instruction.png"))
confirmed_img = ImageTk.PhotoImage(Image.open("img/confirmed.jpg"))
notOnTime_img = ImageTk.PhotoImage(Image.open("img/not_on_time.jpg"))
notreg_img = ImageTk.PhotoImage(Image.open("img/unregistered.jpg"))
redeemed_img = ImageTk.PhotoImage(Image.open("img/redeemed.jpg"))

if service_num == 1:
    header = Image.open("img/welcome1.png")
if service_num == 2:
    header = Image.open("img/welcome2.png")
if service_num == 3:
    header = Image.open("img/welcome3.png")

photo = ImageTk.PhotoImage(header)
headerLabel = Label(image=photo)
headerLabel.image = photo
headerLabel.place(x=64, y=52)

## Body frame ##
body_frame = Frame(ROOT,width=1000,height=600, padx=20, pady=20,bg='white')
body_frame.pack()
body_frame.place(x=550, y=221)

## image labels ##
qrLabel = Label(body_frame, image=qr_img)
confirmedLabel = Label(body_frame, image=confirmed_img)
notLabel = Label(body_frame, image=notOnTime_img)
unregLabel = Label(body_frame, image=notreg_img)
redeemedLabel = Label(body_frame, image=redeemed_img)

## Text labels ##
nameText = Text(body_frame, pady=30, width=32, height=1, font=("Arial", 30))
descText = Text(body_frame, width=32, height=1, font=("Arial", 30))

## Main function ##
exit_loop = False
qrLabel.pack()
while exit_loop == False:
	scan_num = input("Scan QR Code: ")
	if scan_num == "q":
		exit_loop = True
	else:
		if scan_num in attendee_list:
			now = datetime.now()
			name = attendee_list[scan_num][0]
			email = attendee_list[scan_num][1]
			phone = attendee_list[scan_num][2]
			checkIn = attendee_list[scan_num][3]

			## REDEEMED CODE ##
			if scan_num in entered_person:
				print("ERROR: This ticket is already redeemed")
				print("       Ticket# :" + scan_num)
				print("       Name    :" + name)
				print("       Entered :" + entered_person[scan_num])
				sys.stdout.write('\r\a')
				sys.stdout.flush()
				# GUI
				clear_frame(body_frame)
				nameText.delete("1.0", END)
				descText.delete("1.0", END)
				nameText.insert(tk.END, name)
				descText.insert(tk.END,"Redeemed time: " + entered_person[scan_num])
				redeemedLabel.pack()
				nameText.pack()
				descText.pack()

			else:
				## RESERVATION TIME NOT MATCHED ##
				if checkIn != this_service:
						print("Reservation time does not match.")
						sys.stdout.write('\r\a')
						sys.stdout.flush()
						# GUI
						Label(body_frame, image=notOnTime_img, font="bold", pady=20).pack(side=TOP)
						Label(body_frame, text=name, font=("Arial", 19), pady=16).pack()
						Label(body_frame, text="Reserved time: " + checkIn, font=("Arial", 19)).pack()
						clear_frame(body_frame)
						notLabel.pack()
						nameText.delete("1.0", END)
						descText.delete("1.0", END)
						nameText.insert(tk.END, name)
						descText.insert(tk.END, "Reserved time: " + checkIn)
						nameText.pack()
						descText.pack()

				## NEW SUCCESSFUL ENTRY PROCESSING ##
				else:
					print(  name + " " + now.strftime("%H:%M:%S"))
					f.write(name + ", " + email + ", " + phone + ", " + now.strftime("%H:%M:%S") + "\n")
					entered_person[scan_num] = now.strftime("%H:%M:%S")
					# GUI
					clear_frame(body_frame)
					nameText.delete("1.0", END)
					descText.delete("1.0", END)
					nameText.insert(tk.END, name)
					descText.insert(tk.END, "Entered at: " + now.strftime("%H:%M:%S"))
					confirmedLabel.pack()
					nameText.pack()
					descText.pack()
					
		## CODE NOT EXIST ##
		else:
			print("Cannot find ticket #: " + scan_num)
			sys.stdout.write('\r\a')
			sys.stdout.flush()
			# GUI
			clear_frame(body_frame)
			unregLabel.pack()	
